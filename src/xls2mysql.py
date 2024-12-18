# -*- coding: utf-8 -*-
# import pymysql
import os
import argparse
import datetime as dt
import openpyxl
import mysql.connector as mysql
from openpyxl.worksheet.worksheet import Worksheet


def detect_counter_id(template, pos, vertical: bool = False):
    if vertical:
        if pos == 2:
            return 60  # lenta
        elif pos == 3:
            return 61  # central
        elif pos == 4:
            return 65  # door 31
        elif pos == 5:
            return 64  # foodpark
        elif pos == 6:
            return 62  # left atrium
        elif pos == 7:
            return 63  # right atrium
    else:
        if pos == 2 + template:
            return 60  # lenta
        elif pos == 3 + template:
            return 61  # central
        elif pos == 4 + template:
            return 65  # door 31
        elif pos == 5 + template:
            return 64  # foodpark
        elif pos == 6 + template:
            return 62  # left atrium
        elif pos == 7 + template:
            return 63  # right atrium


def record_exists(cursor, date, counter_id):
    sql = "SELECT COUNT(PK_FK) FROM counter_value WHERE PK_FK=%s AND DATE_=%s"
    cursor.execute(sql, (counter_id, date))
    r = cursor.fetchone()
    return r[0] != 0


def write_value_into_mysql(db, cursor, date, counter_id, value):
    if record_exists(cursor, date, counter_id):
        sql = "UPDATE counter_value SET VAL=%s WHERE PK_FK=%s AND DATE_=%s"
        cursor.execute(sql, (value, counter_id, date))
    else:
        sql = "INSERT INTO counter_value (PK_FK, DATE_, VAL) VALUES (%s, %s, %s)"
        cursor.execute(sql, (counter_id, date, value))
    db.commit()


def detect_report_date(reportfilename: str, sheet: Worksheet, row: int = 1, col: int = 4):
    def extract_date(i: int, s: str):
        p1 = s.find('.', i + 7)
        p2 = s.find(' ', i + 7)

        if p1 == -1:
            if p2 == -1:
                result = s[i:]
            else:
                result = s[i:p2]
        elif p2 == -1:
            if p1 == -1:
                result = s[i:]
            else:
                result = s[i:p1]
        else:
            result = s[i:min(p1, p2)]
        return result

    # -----------------------------------------------------------------
    if sheet.cell(row=row, column=col).value is None:  # single day report, date is extracted from filename
        i = 1
        l = len(reportfilename)
        while (i <= l) and (not reportfilename[i].isdigit()):
            i = i + 1
        if i > l: return None

        return extract_date(i, reportfilename)
    else:  # report for period, several dates in row, returned first
        val = sheet.cell(row=row, column=col).value
        if isinstance(val, dt.datetime):
            return val.strftime('%d.%m.%Y')
        else:
            return str(val)


def write_xlsx_to_mysql(file, db, cursor, args):
    if args.verbose:
        print('Prosessing file ' + file + ' ....')

    wb = openpyxl.load_workbook(file)
    sheet = wb['Page 1']

    if hasattr(args, "layout") and isinstance(args.layout, str):
        vertical_layout = str(args.layout).lower().startswith('v')
    else:
        vertical_layout = None
    report_is_singleday = (sheet.cell(row=4, column=5).value is None)

    if vertical_layout == None:
        vertical_layout = report_is_singleday or not (sheet.cell(row=13, column=1).value is None)

    if vertical_layout:
        currow = args.start_row
        while True:
            date = detect_report_date(os.path.basename(file), sheet, row=currow, col=1)
            if date is None: break
            if isinstance(date, str) and date.isalpha(): break
            if len(str(date)) == 10:
                dateformat = '%d.%m.%Y'
            elif len(str(date)) == 8:
                dateformat = '%d.%m.%y'
            else:
                break
            date = dt.datetime.strptime(str(date), dateformat).strftime('%Y-%m-%d')

            for curcol in range(2, 8):
                value = sheet.cell(row=currow, column=curcol).value
                if value is None:
                    continue
                elif isinstance(value, str) and value.isdigit():
                    value = int(value)
                elif not isinstance(value, int):
                    continue
                if value:
                    counter_id = detect_counter_id(int(report_is_singleday), curcol, vertical_layout)
                    write_value_into_mysql(db, cursor, date, counter_id, value)
            if report_is_singleday: break
            currow += 1
    else:               # horizontal layout
        curcol = 4
        while True:
            date = detect_report_date(os.path.basename(file), sheet, row=1, col=curcol)
            if date is None:
                break
            if isinstance(date, str) and date.isalpha(): break
            if len(str(date)) == 10:
                dateformat = '%d.%m.%Y'
            elif len(str(date)) == 8:
                dateformat = '%d.%m.%y'
            else:
                break

            date = dt.datetime.strptime(str(date), dateformat).strftime('%Y-%m-%d')

            for r in range(2, 8 + int(report_is_singleday)):
                value = sheet.cell(row=r, column=curcol).value
                if value is None:
                    continue
                elif isinstance(value, str) and value.isdigit():
                    value = int(value)
                elif not isinstance(value, int):
                    continue
                if value:
                    counter_id = detect_counter_id(int(report_is_singleday), r, vertical_layout)
                    write_value_into_mysql(db, cursor, date, counter_id, value)
            if report_is_singleday: break
            curcol += 1
    # today = dt.today()
    #
    # for i in range(-7, 0):
    #     day = today + tdelta(days=i)
    #     sheet.cell(row=1, column=(11 + i)).value = day.strftime('%d.%m.%y')
    #     sheet.cell(row=5, column=(11 + i)).value = getcounterdata(60, day)
    #     sheet.cell(row=6, column=(11 + i)).value = getcounterdata(62, day)
    #     sheet.cell(row=7, column=(11 + i)).value = getcounterdata(63, day)


def check_arguments():
    def get_date_from_string(str_date: str):
        result = None
        if str_date.find('.') >= 0:  # check length for y and Y!

            if len(str_date) == 8:
                result = dt.strptime(str_date, '%d.%m.%y').date()
            else:
                result = dt.strptime(str_date, '%d.%m.%Y').date()

        elif str_date.find('/') >= 0:

            if len(str_date) == 8:
                result = dt.strptime(str_date, '%d/%m/%y').date()
            else:
                result = dt.strptime(str_date, '%d/%m/%Y').date()

        elif str_date.find('-') >= 0:

            if len(str_date) == 8:
                result = dt.strptime(str_date, '%y-%m-%d').date()
            else:
                result = dt.strptime(str_date, '%Y-%m-%d').date()

        return result

    # анализируем параметры из командной строки

    parser = argparse.ArgumentParser(prog='xls2mysql', description='write data into db from excel report',
                                     fromfile_prefix_chars='@', prefix_chars='-/', usage=__doc__)
    parser.add_argument("--start_row", type=int, default="4", dest="start_row",
                        help="The first row to start data read from")
    parser.add_argument("--verbose", dest="verbose",
                        action='store_true', default=False,
                        help="Verbose output for debugging")
    parser.add_argument("-f", "--force", dest="force",
                        action='store_true', default=False,
                        help="Overwrite data in database even if the record already exists")
    parser.add_argument("reportsdir",  help="directory with reports to put into database ")
    parser.add_argument("--layout", dest="layout",
                        help="Report table layout: horizontal or vertical (default)")
    # parser.add_argument("num", nargs=’ * ’)
    # parser.add_argument("-i", "--interval", type=str, default="days", dest="interval",
    #                     help="Interval sums in cells. Values are: days, weeks, months, years")
    # parser.add_argument("-l", "--layout", type=str, default="", dest="layout",
    #                     help="Report table orientation. Values are: horizontal, vertical")
    # parser.add_argument("-s", "--start", type=str, default="", dest="start",
    #                     help="Report start date. Values are: today, yesterday, any date")
    # parser.add_argument("-e", "--end", type=str, default="", dest="end",
    #                     help="Report end date. Values are: today, yesterday, any date")
    # parser.add_argument("-p", "--period", type=str, default="", dest="period",
    #                     )

    # parser.add_argument("-t", "--test", dest="test",
    #                     action='store_true', default=False,
    #                     help="Test mode: reads from stdin")

    # parser.add_argument("-v", "--verical", dest="vertical",
    #                     action='store_true', default=False,
    #                     help="excel report table layout is vertical")

    # parser.add_argument("-o", "--open_after", default=False, dest="open_after", action='store_true',
    #                     help="Open report file in Excel when ready. Default is False")

    args = parser.parse_args()


    # if no_arguments_specified(args):
    #     # display help
    #     return None
    #
    # for a in vars(args):
    #     if isinstance(a, str):
    #         a = a.lower()
    #
    # today = dt.today().date()
    #
    # if args.start and isinstance(args.start, str):
    #     if args.start == 'yesterday':
    #         args.start = today - tdelta(days=1)
    #     elif args.start == 'today':
    #         args.start = today
    #     else:
    #         args.start = get_date_from_string(args.start)
    #     if args.end and isinstance(args.end, str):
    #         if args.end == 'yesterday':
    #             args.end = today - tdelta(days=1)
    #         elif args.end == 'today':
    #             args.end = today
    #         else:
    #             args.end = get_date_from_string(args.end)
    #
    # elif args.period:
    #     if args.period[:5] == 'lastm':
    #         mr = monthrange(today.year, today.month)
    #         month_ago = today - tdelta(days=mr[1])
    #         args.start = datetime.date(month_ago.year, month_ago.month, 1)
    #         args.end = datetime.date(today.year, today.month, 1) - tdelta(days=1)
    #     elif args.period[:5] == 'lastw':
    #         args.start = today - tdelta(days=7)
    #         while dt.weekday(args.start): args.start = args.start - tdelta(days=1)
    #         args.end = args.start + tdelta(days=6)
    # else:
    #     return None
    #
    # args.interval = args.interval[0]
    # args.layout = args.layout[0]
    #
    # if args.interval not in ('d', 'w', 'm', 'y'): args.interval = 'd'
    # if args.layout not in ('h', 'v'): args.layout = 'h'

    return args


def main():
    # открываем базу
    db = mysql.connect(host='192.168.9.82', user='sa', password='Kristall_123456', database='days')
    cursor = db.cursor()
    # cur.execute("SELECT VERSION()")
    # version = cur.fetchone()
    # print("Database version: {}".format(version[0]))


    args_valid = True

    args = check_arguments()
    reports_processed = False
    if hasattr(args, "reportsdir"):     # single report file or reports directory is specified directly
        if os.path.isfile(args.reportsdir):         # it's a single file
            if os.path.exists(args.reportsdir):     # it exists
                f = os.path.split(args.reportsdir)[1]
                if (not f.startswith('~$')) and (f.endswith('xlsx') or f.endswith('xls')):
                    write_xlsx_to_mysql(args.reportsdir, db, cursor, args)
            reports_processed = True
        else:
            reportsdir = args.reportsdir
    else:
        reportsdir = os.getcwd()
    if not reports_processed:
    # формируем список *.xlsx файлов в каталоге отчетов
        if os.path.isabs(args.reportsdir):
            reportsdir = args.reportsdir
        else:
            reportsdir = os.path.join(os.getcwd(), args.reportsdir)

        if not args_valid: exit(1)
        processed = list()
        for root, dirs, files in os.walk(reportsdir):
            for f in files:
                if f.startswith('~$') or not (f.endswith('xlsx')) or (f in processed): continue
                write_xlsx_to_mysql(os.path.join(root, f), db, cursor, args)
                processed.append(os.path.join(root, f))
                # открываем каждый файл, читаем данные, пишем в базу
        # for f in files: print(f)
        # wb.save(reportfilename)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()
