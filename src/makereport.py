import encodings
from datetime import datetime as dt
from datetime import timedelta as tdelta
import os
import shutil
from pathlib import Path

import openpyxl
import csv
import datetime
import argparse
from calendar import monthrange
from copy import copy
import openpyxl.utils.cell as cellutils
from win32com.client import Dispatch


def no_arguments_specified(args):
    return (args.period in ['day', 'month', 'year', 'hour']) and (not args.start) and (not args.end)


def check_arguments():
    # ! makes sure start and end period dates are specified
    def get_date_from_string(str_date):
        result = None
        if str_date.find('.') >= 0:  # check length for y and Y!

            if len(str_date) == 8:
                result = dt.strptime(str_date, '%d.%m.%y').date()
            else:
                result = dt.strptime(str_date, '%d.%m.%Y').date()

        elif str_date.find('/') >= 0:

            if len(str_date) == 8:
                result = dt.strptime(str_date, '%d/%m/%y').date()
            else:
                result = dt.strptime(str_date, '%d/%m/%Y').date()

        elif str_date.find('-') >= 0:

            if len(str_date) == 8:
                result = dt.strptime(str_date, '%y-%m-%d').date()
            else:
                result = dt.strptime(str_date, '%Y-%m-%d').date()

        return result

    parser = argparse.ArgumentParser(prog='makereport', description='fetch data from devices and create excel report',
                                     fromfile_prefix_chars='@', prefix_chars='-/', usage=__doc__)

    parser.add_argument("-i", "--interval", type=str, default="days", dest="interval",
                        help="Interval sums in cells. Values are: days, weeks, months, years")
    parser.add_argument("-l", "--layout", type=str, default="horizontal", dest="layout",
                        help="Report table orientation. Values are: horizontal, vertical")
    parser.add_argument("-s", "--start", type=str, default="", dest="start",
                        help="Report start date. Values are: today, yesterday, any date")
    parser.add_argument("-e", "--end", type=str, default="", dest="end",
                        help="Report end date. Values are: today, yesterday, any date")
    parser.add_argument("-p", "--period", type=str, default="", dest="period",
                        help="Report interval used instead of start and end dates. Values are: day, week, month, year, lastweek, lastmonth, lastyear, thisweek, thismonth, thisyear")

    parser.add_argument("-a", "--address", type=str, default="", dest="ips",
                        help="Comma separated list of last parts of target devices ip addresses, e.g. 60,61,62")

    parser.add_argument("-r", "--rows", type=str, default="", dest="rows",
                        help="Comma separated list of row numbers for cells corresponding to target devices ip addresses, e.g. 5,6,8")

    parser.add_argument("-c", "--cols", type=str, default="", dest="cols",
                        help="Comma separated list of column numbers for cells corresponding to target devices ip addresses, e.g. 5,6,8")

    parser.add_argument("-t", "--test", dest="test", action='store_true', default=False,
                        help="Test mode: reads from stdin")

    parser.add_argument("-o", "--open_after", dest="open_after", default=False, action='store_true',
                        help="Open report file in Excel when ready. Default is False")

    args = parser.parse_args()

    if no_arguments_specified(args):
        # display help
        return None
    if args.open_after == None: args.open_after = False

    for a in vars(args):
        if isinstance(a, str):
            a = a.lower()

    today = dt.today().date()

    if args.start and isinstance(args.start, str):
        if args.start == 'yesterday':
            args.start = today - tdelta(days=1)
        elif args.start == 'today':
            args.start = today
        else:
            args.start = get_date_from_string(args.start)
        if args.end and isinstance(args.end, str):
            if args.end == 'yesterday':
                args.end = today - tdelta(days=1)
            elif args.end == 'today':
                args.end = today
            else:
                args.end = get_date_from_string(args.end)
        elif not args.end:
            if not args.period:
                args.end = args.start
                args.period = 'day'
        if args.start and args.end and args.start == args.end: args.period = 'day'

    elif args.period:
        if args.period[:5] == 'lastm':
            mr = monthrange(today.year, today.month)
            month_ago = today - tdelta(days=mr[1])
            args.start = datetime.date(month_ago.year, month_ago.month, 1)
            args.end = datetime.date(today.year, today.month, 1) - tdelta(days=1)
        elif args.period[:5] == 'lastw':
            args.start = today - tdelta(days=7)
            while dt.weekday(args.start): args.start = args.start - tdelta(days=1)
            args.end = args.start + tdelta(days=6)
        elif (args.period[:5] == 'lastd') or (args.period[:1] == 'y'):
            args.start = today - tdelta(days=1)
            args.end = args.start
            args.period = 'day'
    else:
        return None

    args.interval = args.interval[0]
    args.layout = args.layout[0]

    if args.interval not in ('d', 'w', 'm', 'y'): args.interval = 'd'
    if args.layout not in ('h', 'v'): args.layout = 'h'

    if args.ips:
        ips = str(args.ips).split(',')
        args.ips = []
        for s in ips:
            args.ips.append(int(s))
    else:
        args.ips = [62, 63, 64]

    if args.layout == 'h':
        args.cols = None
        if args.rows:
            rows = str(args.rows).split(',')
            args.rows = []
            for s in rows:
                args.rows.append(int(s))
        else:
            args.rows = [5, 6, 8]
    else:
        args.rows = None

        if args.cols:
            cols = str(args.cols).split(',')
            args.cols = []
            for s in cols:
                args.cols.append(int(s))
        else:
            args.cols = [2, 3, 4, 5]

    return args


def prepare_blank_report(args, ext: str = '.xlsx'):
    newfilename = ''
    par_dir = Path(__file__).parent.absolute().parent.absolute()
    templatedir = os.path.join(par_dir, "data", "templates")
    reportsdir = os.path.join(par_dir, "data", "reports")

    if args and args.start and args.end:
        create_period_report = args.period != 'day'
        # today = dt.today()
        # yd = today - tdelta(days=1)
        # prevmonday = today - tdelta(days=7)


        if create_period_report:

            if args.layout == 'v':
                templatename = os.path.join(templatedir, 'шаблон отчета за период вертикальный' + ext)
            else:
                templatename = os.path.join(templatedir, 'шаблон отчета за неделю' + ext)

            newfilename = os.path.join(reportsdir,
                                       'отчет за ' + args.start.strftime('%d.%m.%y') + ' - ' + args.end.strftime(
                                           '%d.%m.%y') + ext)
        else:
            templatename = os.path.join(templatedir, 'Шаблон отчета за день' + ext)
            newfilename = os.path.join(reportsdir, 'отчет за ' + args.start.strftime('%d.%m.%y') + ext)

        if os.path.exists(templatename):
            shutil.copy(templatename, newfilename)
    else:  # arguments are not specified

        today = dt.today()
        yd = today - tdelta(days=1)
        prevmonday = today - tdelta(days=7)
        dow = dt.weekday(today)
        create_period_report = dow == 0

        templatedir = os.path.join(os.getcwd(), 'шаблоны')
        reportsdir = os.path.join(os.getcwd(), 'reports')
        if create_period_report:
            templatename = os.path.join(templatedir, 'шаблон отчета за неделю' + ext)
            newfilename = os.path.join(reportsdir, 'отчет за ' + prevmonday.strftime('%d.%m.%y') + ' - ' + yd.strftime(
                '%d.%m.%y') + ext)
        else:
            templatename = os.path.join(templatedir, 'Шаблон отчета за день' + ext)
            if args.period == 'today':
                newfilename = os.path.join(reportsdir, 'отчет за ' + today.strftime('%d.%m.%y') + ext)
            elif args.period[0] == 'y':
                newfilename = os.path.join(reportsdir, 'отчет за ' + today.strftime('%d.%m.%y') + ext)
            else:
                newfilename = os.path.join(reportsdir, 'отчет за ' + args.start + ext)

        if os.path.exists(templatename):
            shutil.copy(templatename, newfilename)
    return newfilename


def getcounterdata(number, date: dt):
    result = 0
    ip = '192.168.5.' + str(number)
    datefrom = date.strftime('%d/%m/%y')

    timeto = '22:00:00.00'
    timefrom = '08:00:00.00'
    outfile = os.path.join(os.getcwd(), str(number) + '.csv')
    searchword = 'Вход'

    query = ' ip=' + ip + ' datefrom=' + datefrom + ' dateto=' + datefrom + ' timefrom=' + timefrom + \
            ' timeto=' + timeto + ' outfile=' + outfile

    if (not os.system('GetReport.exe' + query)) and os.path.exists(outfile):
        with open(outfile, 'r', encoding='utf-8') as srcfile, open(outfile + '.tmp', 'w', encoding='utf-8') as dstfile:
            i = 0
            for line in srcfile:
                if i < 2:
                    i += 1
                    continue
                dstfile.write(line)
                i += 1

        os.remove(outfile)
        os.rename(outfile + '.tmp', outfile)

        with open(outfile, 'r', newline='', encoding='utf-8') as csvfile:
            dialect = csv.Dialect
            dialect.delimiter = ';'
            dialect.quoting = csv.QUOTE_MINIMAL
            dialect.doublequote = True
            dialect.escapechar = None
            dialect.lineterminator = '\r\n'
            dialect.skipinitialspace = False
            dialect.strict = False
            dialect.quotechar = '"'

            filedata = csv.DictReader(csvfile, dialect=dialect)
            result = 0
            for row in filedata:
                if (row.get('Текст предупреждения') == 'Вход') or (row.get('Тип прохода') == 'Вход'):
                    result += 1
    return result


def fill_in_dayly_report_data(args, reportfilename):
    if args and args.start:
        daystart = args.start
    elif args.period == 'today':
        daystart = dt.today()
    elif args.period[0] == 'y':
        daystart = dt.today() - tdelta(days=1)
    else:
        return

    wb = openpyxl.load_workbook(reportfilename)
    sheet = wb['Page 1']

    sheet['D2'] = daystart.strftime('%d.%m.%y')
    sheet['C12'] = dt.today().strftime('%d.%m.%y')
    sheet['D6'] = getcounterdata(64, daystart)
    sheet['D7'] = getcounterdata(62, daystart)
    sheet['D8'] = getcounterdata(63, daystart)
    wb.save(reportfilename)


def fill_in_period_report_data(args, reportfilename):
    if not args: return

    first_data_column = 4
    first_data_row = 4
    wb = openpyxl.load_workbook(reportfilename)
    sheet = wb['Page 1']
    today = dt.today()

    if args.start and args.end:
        delta = args.start - args.end
        dayend = args.end
    else:
        delta = -7
        dayend = today

    if args.layout == 'h':
        toprow = 1
        endrow = 10
        columns_with_data__in_horizontal_template = 7

        sheet.cell(1, first_data_column - delta.days + 1).value = 'Итог'

        first_data_column_letter = cellutils.get_column_letter(first_data_column)
        last_data_column_letter = cellutils.get_column_letter(first_data_column - delta.days)

        if delta.days > (1 - columns_with_data__in_horizontal_template):
            sheet.delete_cols(first_data_column - delta.days + 1,
                              columns_with_data__in_horizontal_template + delta.days - 1)
        elif delta.days < (1 - columns_with_data__in_horizontal_template):
            sheet.insert_cols(first_data_column + columns_with_data__in_horizontal_template - 1,
                              columns_with_data__in_horizontal_template + delta.days - 1)

        for i in range(delta.days, 1):
            day = dayend + tdelta(days=i)
            col = i - delta.days + first_data_column
            sheet.cell(row=1, column=col).value = day.strftime('%d.%m.%y')
            if not args.ips or not args.rows: continue
            for j, ip in enumerate(args.ips):
                sheet.cell(row=args.rows[j], column=col).value = getcounterdata(ip, day)
            #            if i > delta.days:
            for row in range(toprow, endrow):
                src_cell = sheet.cell(row, first_data_column)
                dst_cell = sheet.cell(row, col)

                total_cell = sheet.cell(row, first_data_column - delta.days + 1)
                if row == toprow:
                    total_cell.value = 'Итог'
                else:
                    total_cell.value = "=SUM(" + first_data_column_letter + str(
                        row) + ":" + last_data_column_letter + str(row) + ")"

                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = copy(src_cell.number_format)
                    dst_cell.alignment = copy(src_cell.alignment)

    else:
        for i in range(delta.days, 2):
            day = dayend + tdelta(days=i)
            row = i - delta.days + first_data_row
            sheet.cell(row=row, column=1).value = day.strftime('%d.%m.%y')
            if not args.ips or not args.rows: continue
            for j, ip in enumerate(args.ips):
                sheet.cell(row=args.rows[j], column=col).value = getcounterdata(ip, day)

    if 'us-ascii' not in encodings.aliases.aliases:
        encodings.aliases.aliases['us-ascii'] = 'ascii'
    wb.save(reportfilename)


def main():
    today = dt.today()
    dow = dt.weekday(today)
    args = check_arguments()

    reportfilename = prepare_blank_report(args)

    if os.path.exists(reportfilename):
        if (args and (args.period != 'day') and (args.period != 'today')) or ((not args) and (dow == 0)):
            fill_in_period_report_data(args, reportfilename)

        else:
            fill_in_dayly_report_data(args, reportfilename)

    if args:
        resume_file = os.path.join(os.environ['USERPROFILE'], 'Documents', 'RESUME.XLW')
        if os.path.exists(resume_file): os.remove(resume_file)
        xl = Dispatch("Excel.Application")
        xl.Visible = True  # otherwise excel is hidden
        # xl.Save()
        if args.open_after:
            xl.Workbooks.Open(reportfilename)


if __name__ == '__main__':
    main()
