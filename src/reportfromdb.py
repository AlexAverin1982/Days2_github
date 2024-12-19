#! /usr/bin/env python3
"""fetch data from db and create excel report
usage: reportfromdb [parameters]
parameters:
-id, --interval=day: table cells will contain daily sums"""
import datetime

# from datetime import date
import os
from datetime import datetime as dt
from datetime import timedelta as tdelta
from pathlib import Path

import openpyxl
# from openpyxl.cell import Cell, MergedCell
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from calendar import monthrange

# from openpyxl.utils import FORMULAE
# from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.styles import Font, Alignment, Border, Side

# import mysql.connector as mysql
import pymysql
import pymysql.cursors
import argparse
import shutil
from win32com.client import Dispatch


# from PyQt5.QtWidgets import (QWidget, QPushButton, QLineEdit, QInputDialog, QApplication)
# from openpyxl.worksheet.worksheet import Worksheet


def test(args) -> None:
    # выводим результаты парсинга аргументов
    print(vars(args))


def single_day_report(args) -> bool:
    return (
            (str(args.start).find("day") >= 0)
            or (str(args.end).find("day") >= 0)
            or (str(args.period).find("day") >= 0)
    )


def prepare_blank_report(args):
    horizontal_layout = str(args.layout).startswith("h")
    # today = dt.today().date()
    #
    # # yd = (today - tdelta(days=1))
    # prevmonday = (today - tdelta(days=7))
    # while dt.weekday(prevmonday): prevmonday = (prevmonday - tdelta(days=1))

    # if create_period_report:
    #     if not args.start:
    #         args.start = prevmonday
    #         args.start = args.start
    #         args.end = args.start + tdelta(days=6)    # default period report is for last week
    #         args.end = args.end
    #     elif not args.end:
    #         if hasattr(args, "period"):
    #             if str(args.period).lower().startswith("mon"):
    #                 args.end = args.start + tdelta(days=31)
    #                 while args.start.month != args.end.month: args.end = args.end - tdelta(days=1)
    #         else:
    #             args.end = args.start + tdelta(days=6)
    #         args.end = args.end
    # else:
    #     if not args.start:
    #         args.start = args.end = yd
    #         args.end = args.end
    #         args.end = args.end

    par_dir = Path(__file__).parent.absolute().parent.absolute()
    templatedir = os.path.join(par_dir, "data", "templates")
    reportsdir = os.path.join(par_dir, "data", "reports")

    if not single_day_report(args):
        if horizontal_layout:
            templatename = os.path.join(templatedir, "шаблон отчета за период.xlsx")
        else:
            templatename = os.path.join(
                templatedir, "шаблон отчета за период вертикальный.xlsx"
            )
        newfilename = os.path.join(
            reportsdir,
            "отчет за "
            + args.start.strftime("%d.%m.%y")
            + " - "
            + args.end.strftime("%d.%m.%y")
            + ".xlsx",
        )
    else:
        templatename = os.path.join(templatedir, "Шаблон отчета за день.xlsx")
        newfilename = os.path.join(
            reportsdir, "отчет за " + args.start.strftime("%d.%m.%y") + ".xlsx"
        )

    if os.path.exists(templatename):
        shutil.copy(templatename, newfilename)

    return newfilename


def get_data_from_db(cursor, date_start, date_end):
    sql = "SELECT NAME, DATE_, VAL FROM counter_value JOIN counter ON PK=PK_FK " + \
          f"WHERE DATE_ BETWEEN '{date_start}' AND '{date_end}' ORDER BY DATE_, DISPLAYORDER"
    cursor.execute(sql)
    return cursor.fetchall()


def fill_in_dayly_report_data(cursor, args, reportfilename):
    wb = openpyxl.load_workbook(reportfilename)
    sheet = wb["Page 1"]
    yd = (dt.today() - tdelta(days=1)).date()
    if not args.start:
        report_date = yd
    else:
        report_date = args.start

    records = get_data_from_db(cursor, report_date, report_date + tdelta(days=1))

    start_row = 3
    for rec in records:
        cell = "D" + str(start_row)
        val = rec["VAL"]
        if val:
            sheet[cell] = val
        else:
            sheet[cell] = 0
        start_row = start_row + 1

    wb.save(reportfilename)


def set_cell_value_and_style(
        cell: Cell, value, no_borders: bool = False, layout: str = "h"
):
    border_side = Side(border_style="thin")
    square_border = Border(
        top=border_side, right=border_side, bottom=border_side, left=border_side
    )
    align_center = Alignment(horizontal="center", vertical="center")
    font = Font(name="Arial", size=14)

    cell.value = value
    if no_borders:
        cell.border = None
    else:
        cell.border = square_border
    cell.alignment = align_center
    cell.font = font

    cell.number_format = "### ### ###"

    if (layout == "h") and (cell.row in [6, 7, 8]):
        cell.fill = PatternFill(
            start_color="DBDBDB", end_color="DBDBDB", fill_type="solid"
        )

    return cell


def fill_in_period_report_data(cursor, args, reportfilename):
    wb = openpyxl.load_workbook(reportfilename)
    sheet = wb["Page 1"]  # read from ini

    today = dt.today().date()  # setting up period boundaries
    # # yd = (today - tdelta(days=1)).date()
    # prevmonday = today - tdelta(days=7)
    #
    # while dt.weekday(prevmonday): prevmonday = prevmonday - tdelta(days=1)

    records = get_data_from_db(
        cursor, args.start, args.end + tdelta(days=2)
    )  # obtaining data from db
    record_length = 6  # FIXME: read from ini, what is it, number of counters?
    total_days_count = (len(records)) // record_length

    if total_days_count < ((args.end - args.start).days + 1):
        total_days_count = (args.end - args.start).days + 1

    if args.layout == "h":
        # -----------------------------------horizontal layout --------------------------------------------------
        first_day_col = 4  # read from ini             # table settings
        last_record_row = 12  # read from ini
        first_column_letter = "D"  # used in formulae
        last_column_letter = "D"

        days_count = 0  # loop counters
        cur_date = args.start

        for column in sheet.iter_cols(
                min_row=0,
                min_col=first_day_col,
                max_row=last_record_row,
                max_col=total_days_count + first_day_col,
        ):
            set_cell_value_and_style(
                column[0], cur_date.strftime("%d.%m.%Y"), layout=args.layout
            )

            if column[0].col_idx == first_day_col:
                first_column_letter = column[0].column_letter
            else:
                last_column_letter = column[0].column_letter

            row = 1

            day_records = records[
                          days_count * record_length: (days_count + 1) * record_length
                          ]  # getting data slice for  current period
            for rec in day_records:
                if rec["VAL"]:
                    val = rec["VAL"]
                else:
                    val = 0
                set_cell_value_and_style(column[row], val, layout=args.layout)
                row = row + 1

            s = (
                    "=SUM("
                    + column[0].column_letter
                    + str(record_length)
                    + ":"
                    + column[0].column_letter
                    + str(record_length + 1)
                    + ")"
            )
            set_cell_value_and_style(column[row], s, layout=args.layout)
            s = (
                    "=SUM("
                    + column[0].column_letter
                    + "2:"
                    + column[0].column_letter
                    + str(record_length - 1)
                    + ")"
            )
            set_cell_value_and_style(column[row + 1], s, layout=args.layout)
            sheet.merge_cells(
                column[0].column_letter
                + str(row + 2)
                + ":"
                + column[0].column_letter
                + str(row + 3)
            )
            days_count = days_count + 1

            if cur_date == args.end:
                break
            else:
                cur_date = cur_date + tdelta(days=1)
        #  --------------------------------- end of loop ----------------
        ldc = total_days_count + first_day_col - 1

        for column in sheet.iter_cols(
                min_row=1, min_col=ldc, max_row=last_record_row, max_col=ldc
        ):
            set_cell_value_and_style(column[0], "Итог", layout=args.layout)
            for i in range(1, record_length + 3):
                s = (
                        "=SUM("
                        + first_column_letter
                        + str(i + 1)
                        + ":"
                        + last_column_letter
                        + str(i + 1)
                        + ")"
                )
                set_cell_value_and_style(column[i], s, layout=args.layout)

        sheet.merge_cells(
            column[0].column_letter
            + str(record_length + 3)
            + ":"
            + column[0].column_letter
            + str(record_length + 4)
        )

        for column in sheet.iter_cols(min_col=first_day_col, max_col=ldc):
            sheet.column_dimensions[column[0].column_letter].width = 16

        set_cell_value_and_style(
            sheet.cell(row=12, column=1),
            "Дата построения отчета: " + today.strftime("%d.%m.%Y"),
            no_borders=True,
            layout=args.layout,
        )
    else:
        # -----------------------------------vertical layout -------------------------------------------------------
        first_day_col = 4  # TODO: read from ini             # table settings
        first_day_row = 4
        last_record_row = first_day_row + total_days_count  # read from ini
        # first_record_col_letter = "B"
        # last_record_col_letter = "E"
        total_column = 8
        first_column_letter = "A"  # used in formulae
        last_column_letter = "D"

        days_count = 0  # loop counters
        cur_date = args.start
        for row in sheet.iter_rows(
                min_row=first_day_row, min_col=0, max_row=last_record_row, max_col=9
        ):
            set_cell_value_and_style(
                row[0], cur_date.strftime("%d.%m.%Y"), layout=args.layout
            )

            # if row[0].col_idx == first_day_col:
            #     first_row_letter = row[0].row_letter
            # else:
            #     last_row_letter = row[0].row_letter

            col = 1
            day_records = records[
                          days_count * record_length: (days_count + 1) * record_length
                          ]  # getting data slice for  current period

            for rec in day_records:
                if rec["VAL"]:
                    val = rec["VAL"]
                else:
                    val = 0
                if rec["NAME"] == "Lenta":
                    col = 1
                elif rec["NAME"] == "Central":
                    col = 2
                elif rec["NAME"] == "Door31":
                    col = 3
                elif rec["NAME"] == "FoodPark":
                    col = 4
                elif rec["NAME"] == "LeftAtrium":
                    col = 5
                elif rec["NAME"] == "RightAtrium":
                    col = 6
                set_cell_value_and_style(row[col], val, layout=args.layout)
                # col = col + 1
            col = 7
            s = (
                    "=SUM(F"
                    + str(first_day_row + days_count)
                    + ":G"
                    + str(first_day_row + days_count)
                    + ")"
            )
            set_cell_value_and_style(
                row[col], s, layout=args.layout
            )  # second floor total
            s = (
                    "=SUM(B"
                    + str(first_day_row + days_count)
                    + ":E"
                    + str(first_day_row + days_count)
                    + ")"
            )
            set_cell_value_and_style(
                row[total_column], s, layout=args.layout
            )  # first floor total
            # sheet.merge_cells(row[0].row_letter + str(row + 2) + ':' + row[0].row_letter + str(row + 3))
            days_count = days_count + 1

            if cur_date == args.end + tdelta(days=1):
                break
            else:
                cur_date = cur_date + tdelta(days=1)
        #  --------------------------------- end of loop ----------------
        ldr = total_days_count + first_day_row

        for col in sheet.iter_cols(min_row=ldr, min_col=0, max_row=ldr, max_col=9):
            if col[0].col_idx == 1:
                set_cell_value_and_style(col[0], "Итог", layout=args.layout)
            else:
                column_letter = col[0].column_letter
                s = (
                        "=SUM("
                        + column_letter
                        + str(first_day_row)
                        + ":"
                        + column_letter
                        + str(ldr - 1)
                        + ")"
                )
                set_cell_value_and_style(col[0], s, layout=args.layout)

        set_cell_value_and_style(
            sheet.cell(row=1, column=1),
            "Отчет по посещаемости ТРК Ясень за: "
            + args.start.strftime("%d.%m.%Y")
            + " - "
            + args.end.strftime("%d.%m.%Y"),
            no_borders=True,
            layout=args.layout,
        )
        sheet.merge_cells(
            start_row=ldr + 3, start_column=1, end_row=ldr + 3, end_column=4
        )
        set_cell_value_and_style(
            sheet.cell(row=ldr + 3, column=1),
            "Дата построения отчета: " + today.strftime("%d.%m.%Y"),
            no_borders=True,
            layout=args.layout,
        )
    wb.save(reportfilename)


def no_arguments_specified(args) -> bool:
    return (
            (args.period in ["day", "month", "year", "hour"])
            and (not args.start)
            and (not args.end)
    )


def check_arguments():
    def get_date_from_string(str_date: str):
        result = None
        if str_date.find(".") >= 0:  # check length for y and Y!

            if len(str_date) == 8:
                result = dt.strptime(str_date, "%d.%m.%y").date()
            else:
                result = dt.strptime(str_date, "%d.%m.%Y").date()

        elif str_date.find("/") >= 0:

            if len(str_date) == 8:
                result = dt.strptime(str_date, "%d/%m/%y").date()
            else:
                result = dt.strptime(str_date, "%d/%m/%Y").date()

        elif str_date.find("-") >= 0:

            if len(str_date) == 8:
                result = dt.strptime(str_date, "%y-%m-%d").date()
            else:
                result = dt.strptime(str_date, "%Y-%m-%d").date()

        return result

    # анализируем параметры из командной строки

    parser = argparse.ArgumentParser(
        prog="reportfromdb",
        description="fetch data from db and create excel report",
        fromfile_prefix_chars="@",
        prefix_chars="-/",
        usage=__doc__,
    )
    # parser.add_argument("num", nargs=’ * ’)
    parser.add_argument(
        "-i",
        "--interval",
        type=str,
        default="days",
        dest="interval",
        help="Interval sums in cells. Values are: days, weeks, months, years",
    )
    parser.add_argument(
        "-l",
        "--layout",
        type=str,
        default="horizontal",
        dest="layout",
        help="Report table orientation. Values are: horizontal, vertical",
    )
    parser.add_argument(
        "-s",
        "--start",
        type=str,
        default="",
        dest="start",
        help="Report start date. Values are: today, yesterday, any date",
    )
    parser.add_argument(
        "-e",
        "--end",
        type=str,
        default="",
        dest="end",
        help="Report end date. Values are: today, yesterday, any date",
    )
    parser.add_argument(
        "-p",
        "--period",
        type=str,
        default="",
        dest="period",
        help="Report interval used instead of start and end dates. " +
             "Values are: day, week, month, year, lastweek, lastmonth, lastyear, thisweek, thismonth, thisyear",
    )

    parser.add_argument(
        "-t",
        "--test",
        dest="test",
        action="store_true",
        default=False,
        help="Test mode: reads from stdin",
    )

    parser.add_argument(
        "-o",
        "--open_after",
        default=False,
        dest="open_after",
        action="store_true",
        help="Open report file in Excel when ready. Default is False",
    )

    parser.add_argument(
        "-y",
        "--year",
        default=dt.today().date().year,
        dest="year",
        help="The year of report",
    )

    args = parser.parse_args()

    if no_arguments_specified(args):
        parser.usage()
        # display help
        return None

    for a in vars(args):
        if isinstance(a, str):
            a = a.lower()

    today = dt.today().date()
    yd = today - tdelta(days=1)
    prevmonday = today - tdelta(days=7)
    while dt.weekday(prevmonday):
        prevmonday = prevmonday - tdelta(days=1)

    # create_period_report = not single_day_report(args)

    # horizontal_layout = hasattr(args, "layout") and (str(args.layout).startswith('h'))

    if not single_day_report(args):
        if not args.start:
            args.start = prevmonday
            args.end = args.start + tdelta(
                days=6
            )  # default period report is for last week
            args.end = args.end
            # elif not args.end:
        #     args.end = args.start + tdelta(days=6)
    elif not args.start:
        args.start = args.end = yd

    if args.start and isinstance(args.start, str):
        if args.start == "yesterday":
            args.start = today - tdelta(days=1)
        elif args.start == "today":
            args.start = today
        # TODO: more word variants like lastmonday
        else:
            args.start = get_date_from_string(args.start)
        if args.end and isinstance(args.end, str):
            if args.end == "yesterday":
                args.end = today - tdelta(days=1)
            elif args.end == "today":
                args.end = today
            else:
                args.end = get_date_from_string(args.end)

    if args.period:

        args.period = str(args.period).lower()
        period_data = [
            (1, "ja", 31),
            (2, "f", 28),
            (3, "mar", 31),
            (4, "ap", 30),
            (5, "may", 31),
            (6, "jun", 30),
            (7, "jul", 31),
            (8, "au", 31),
            (9, "s", 30),
            (10, "o", 31),
            (11, "n", 30),
            (12, "d", 31),
        ]
        period_detected = False

        for n, m, e in period_data:
            period_detected = args.period.startswith(m)
            if period_detected:
                args.start = datetime.date(args.year, n, 1)
                if args.start > today:
                    args.year = args.year - 1
                    args.start = datetime.date(args.year, n, 1)
                if (n == 2) and (args.year % 4):
                    e = 29
                args.end = datetime.date(args.year, n, e)
                break
        if not period_detected:
            if args.period[:5] == "lastm":
                mr = monthrange(today.year, today.month)
                month_ago = today - tdelta(days=mr[1])
                args.start = datetime.date(month_ago.year, month_ago.month, 1)
                args.end = datetime.date(today.year, today.month, 1) - tdelta(days=1)
            elif args.period[:5] == "lastw":
                args.start = today - tdelta(days=7)
                while dt.weekday(args.start):
                    args.start = args.start - tdelta(days=1)
                args.end = args.start + tdelta(days=6)
            elif str(args.period).lower().startswith("mon"):
                args.end = args.start + tdelta(days=31)
                while args.start.month != args.end.month:
                    args.end = args.end - tdelta(days=1)
            elif str(args.period).lower().startswith("mar"):
                if today.month < 3:
                    args.start = datetime.date(today.year - 1, 3, 1)
                    args.end = datetime.date(today.year - 1, 3, 31)
            elif str(args.period).lower().startswith("apr"):
                if today.month < 4:
                    args.start = datetime.date(today.year - 1, 4, 1)
                    args.end = datetime.date(today.year - 1, 4, 30)
                else:
                    args.start = datetime.date(today.year, 4, 1)
                    args.end = datetime.date(today.year, 4, 30)
            elif str(args.period).lower().startswith("may"):
                if today.month < 5:
                    args.start = datetime.date(today.year - 1, 5, 1)
                    args.end = datetime.date(today.year - 1, 5, 31)
                else:
                    args.start = datetime.date(today.year, 5, 1)
                    args.end = datetime.date(today.year, 5, 31)
            elif str(args.period).lower().startswith("jun"):
                if today.month < 6:
                    args.start = datetime.date(today.year - 1, 6, 1)
                    args.end = datetime.date(today.year - 1, 6, 30)
                else:
                    args.start = datetime.date(today.year, 6, 1)
                    args.end = datetime.date(today.year, 6, 30)
            elif str(args.period).lower().startswith("jul"):
                if today.month < 7:
                    args.start = datetime.date(today.year - 1, 7, 1)
                    args.end = datetime.date(today.year - 1, 7, 31)
                else:
                    args.start = datetime.date(today.year, 7, 1)
                    args.end = datetime.date(today.year, 7, 31)
            else:
                args.start = datetime.date(today.year, 3, 1)
                args.end = datetime.date(today.year, 3, 31)
                # args.end = args.start + tdelta(days=31)
    else:
        return None

    args.interval = args.interval[0]
    args.layout = args.layout[0]

    if args.interval not in ("d", "w", "m", "y"):
        args.interval = "d"
    if args.layout not in ("h", "v"):
        args.layout = "h"

    return args


def main():
    args = check_arguments()
    # to this moment, period borders in args must be specified explicitly! no more recalculations!
    if args.test:
        test(args)
    db = pymysql.connect(
        host="192.168.9.82",
        user="sa",
        password="Kristall_123456",
        database="days",
        cursorclass=pymysql.cursors.DictCursor,
    )
    cursor = db.cursor()

    reportfilename = prepare_blank_report(args)

    if single_day_report(args):
        fill_in_dayly_report_data(cursor, args, reportfilename)
    else:
        fill_in_period_report_data(cursor, args, reportfilename)

    if args.open_after:
        if os.path.exists(
                os.path.join(os.environ["USERPROFILE"], "Documents", "RESUME.XLW")
        ):
            os.remove(
                os.path.join(os.environ["USERPROFILE"], "Documents", "RESUME.XLW")
            )
        xl = Dispatch("Excel.Application")
        xl.Visible = True  # otherwise excel is hidden
        xl.Workbooks.Open(reportfilename)
        # xl.Save()
        # cur.execute("SELECT VERSION()")
        # version = cur.fetchone()
        # print("Database version: {}".format(version[0]))


# Press the green button in the gutter to run the script.
if __name__ == "__main__":
    main()
